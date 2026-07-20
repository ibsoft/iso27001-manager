import csv
import io
import os
import uuid
from datetime import datetime
from werkzeug.datastructures import FileStorage

from flask import Blueprint, render_template, redirect, url_for, flash, request, Response, current_app
from flask_login import login_required, current_user
from flask_babel import gettext as _
from app.extensions import db
from app.models.asset import Asset
from app.models.user import User
from app.models.audit_log import AuditLog
from app.forms import AssetForm
from sqlalchemy import or_
from app.utils.decorators import permission_required, admin_required
from app.utils.pagination import paginate

assets_bp = Blueprint("assets", __name__)

ALLOWED_PICTURE_EXTENSIONS = {"png", "jpg", "jpeg", "gif", "webp"}

def _allowed_picture(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_PICTURE_EXTENSIONS

def _save_picture(file_storage):
    ext = file_storage.filename.rsplit(".", 1)[1].lower()
    filename = f"asset_{uuid.uuid4().hex}.{ext}"
    upload_dir = os.path.join(current_app.config["UPLOAD_FOLDER"], "assets")
    os.makedirs(upload_dir, exist_ok=True)
    file_storage.save(os.path.join(upload_dir, filename))
    return filename

def _delete_picture(filename):
    if filename:
        path = os.path.join(current_app.config["UPLOAD_FOLDER"], "assets", filename)
        if os.path.exists(path):
            os.remove(path)


@assets_bp.route("/")
@login_required
@permission_required("menu_assets")
def list_assets():
    classification = request.args.get("classification")
    asset_type = request.args.get("asset_type")
    status = request.args.get("status")
    search = request.args.get("search", "")

    query = Asset.query
    if classification:
        query = query.filter_by(classification=classification)
    if asset_type:
        query = query.filter_by(asset_type=asset_type)
    if status:
        query = query.filter_by(status=status)
    if search:
        query = query.filter(
            or_(Asset.name.ilike(f"%{search}%"), Asset.barcode == search)
        )

    assets = paginate(query.order_by(Asset.name))
    return render_template("assets/list.html", assets=assets)


@assets_bp.route("/export")
@login_required
def export_assets():
    fmt = request.args.get("format", "csv")
    classification = request.args.get("classification")
    asset_type = request.args.get("asset_type")
    status = request.args.get("status")
    search = request.args.get("search", "")

    query = Asset.query
    if classification:
        query = query.filter_by(classification=classification)
    if asset_type:
        query = query.filter_by(asset_type=asset_type)
    if status:
        query = query.filter_by(status=status)
    if search:
        query = query.filter(
            or_(Asset.name.ilike(f"%{search}%"), Asset.barcode == search)
        )
    assets = query.order_by(Asset.name).all()

    headers = [_("Name"), _("Serial Number"), _("Description"), _("Type"),
               _("Classification"), _("Criticality"), _("Status"), _("Location"),
               _("Barcode / QR Code"), _("Owner"), _("Retention Period"), _("Notes"), _("Created At"), _("Updated At")]

    if fmt == "xlsx":
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = _("Assets")
            ws.append(headers)
            for a in assets:
                ws.append([a.name, a.serial_number, a.description, a.asset_type,
                          a.classification, a.criticality, a.status, a.location,
                          a.barcode or "",
                          f"{a.owner.first_name} {a.owner.last_name}" if a.owner else "",
                          a.retention_period, a.notes,
                          a.created_at.strftime("%Y-%m-%d %H:%M") if a.created_at else "",
                          a.updated_at.strftime("%Y-%m-%d %H:%M") if a.updated_at else ""])
            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            return Response(bio.getvalue(),
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=assets_{datetime.now().strftime('%Y%m%d')}.xlsx"})
        except Exception as e:
            flash(_("Export failed: %(error)s", error=str(e)), "danger")
            return redirect(url_for("assets.list_assets"))

    si = io.StringIO()
    writer = csv.writer(si)
    writer.writerow(headers)
    for a in assets:
        writer.writerow([a.name, a.serial_number, a.description, a.asset_type,
                        a.classification, a.criticality, a.status, a.location,
                        a.barcode or "",
                        f"{a.owner.first_name} {a.owner.last_name}" if a.owner else "",
                        a.retention_period, a.notes,
                        a.created_at.strftime("%Y-%m-%d %H:%M") if a.created_at else "",
                        a.updated_at.strftime("%Y-%m-%d %H:%M") if a.updated_at else ""])
    output = si.getvalue()
    si.close()
    return Response(output, mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=assets_{datetime.now().strftime('%Y%m%d')}.csv"})


@assets_bp.route("/import", methods=["GET", "POST"])
@login_required
@permission_required("asset_create")
def import_assets():
    results = {"created": 0, "skipped": 0, "errors": []}
    if request.method == "POST":
        file = request.files.get("file")
        if not file or file.filename == "":
            flash(_("No file selected."), "danger")
            return render_template("assets/import.html", results=results)

        ext = os.path.splitext(file.filename)[1].lower()
        rows = []
        try:
            if ext == ".csv":
                content = file.read().decode("utf-8-sig")
                reader = csv.DictReader(io.StringIO(content))
                rows = list(reader)
            elif ext in (".xlsx", ".xls"):
                from openpyxl import load_workbook
                wb = load_workbook(file, read_only=True, data_only=True)
                ws = wb.active
                header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(header_row, [str(v) if v is not None else "" for v in row])))
            else:
                flash(_("Unsupported file format. Use CSV or XLSX."), "danger")
                return render_template("assets/import.html", results=results)
        except Exception as e:
            flash(_("Error reading file: %(error)s", error=str(e)), "danger")
            return render_template("assets/import.html", results=results)

        if not rows:
            flash(_("No data found in file."), "warning")
            return render_template("assets/import.html", results=results)

        owner_cache = {u.id: u for u in User.query.filter_by(is_active=True).all()}
        owner_name_map = {f"{u.first_name} {u.last_name}".lower(): u.id for u in owner_cache.values()}

        for idx, row in enumerate(rows, start=2):
            name = (row.get(_("Name")) or row.get("Name") or "").strip()
            if not name:
                results["skipped"] += 1
                continue

            try:
                asset = Asset(name=name)
                asset.serial_number = (row.get(_("Serial Number")) or row.get("Serial Number") or "").strip()
                asset.description = (row.get(_("Description")) or row.get("Description") or "").strip()
                asset.asset_type = (row.get(_("Type")) or row.get("Type") or "").strip().lower()
                asset.classification = (row.get(_("Classification")) or row.get("Classification") or "internal").strip().lower()
                asset.criticality = (row.get(_("Criticality")) or row.get("Criticality") or "medium").strip().lower()
                asset.status = (row.get(_("Status")) or row.get("Status") or "active").strip().lower()
                asset.location = (row.get(_("Location")) or row.get("Location") or "").strip()
                asset.retention_period = (row.get(_("Retention Period")) or row.get("Retention Period") or "").strip()
                asset.notes = (row.get(_("Notes")) or row.get("Notes") or "").strip()
                asset.barcode = (row.get(_("Barcode / QR Code")) or row.get("Barcode / QR Code") or "").strip()

                owner_name = (row.get(_("Owner")) or row.get("Owner") or "").strip()
                if owner_name:
                    oid = owner_name_map.get(owner_name.lower())
                    if oid:
                        asset.owner_id = oid

                db.session.add(asset)
                results["created"] += 1
            except Exception as e:
                results["errors"].append(_("Row %(row)d: %(error)s", row=idx, error=str(e)))
                results["skipped"] += 1

        db.session.commit()
        if results["created"]:
            _log_audit_action(f"Imported {results['created']} assets from file: {file.filename}")
            flash(_("Imported %(count)d assets.", count=results["created"]), "success")
        if results["errors"]:
            for err in results["errors"][:5]:
                flash(err, "warning")

    return render_template("assets/import.html", results=results)


@assets_bp.route("/new", methods=["GET", "POST"])
@login_required
@permission_required("asset_create")
def new_asset():
    form = AssetForm()
    form.owner_id.choices = [(0, _("Unassigned"))] + [(u.id, f"{u.first_name} {u.last_name}") for u in User.query.filter_by(is_active=True).all()]

    if form.validate_on_submit():
        asset = Asset()
        form.populate_obj(asset)
        if form.owner_id.data == 0:
            asset.owner_id = None
        if isinstance(form.picture.data, FileStorage):
            if _allowed_picture(form.picture.data.filename):
                asset.picture = _save_picture(form.picture.data)
            else:
                flash(_("Invalid picture format. Allowed: PNG, JPG, JPEG, GIF, WebP"), "warning")
        db.session.add(asset)
        db.session.commit()
        _log_audit(f"Created asset: {asset.name}")
        flash(_("Asset created successfully."), "success")
        return redirect(url_for("assets.view_asset", asset_id=asset.id))

    return render_template("assets/form.html", form=form, title=_("New Asset"))


@assets_bp.route("/<int:asset_id>")
@login_required
def view_asset(asset_id):
    asset = Asset.query.get_or_404(asset_id)
    return render_template("assets/view.html", asset=asset)


@assets_bp.route("/<int:asset_id>/edit", methods=["GET", "POST"])
@login_required
@permission_required("asset_edit")
def edit_asset(asset_id):
    asset = Asset.query.get_or_404(asset_id)
    form = AssetForm(obj=asset)
    form.owner_id.choices = [(0, _("Unassigned"))] + [(u.id, f"{u.first_name} {u.last_name}") for u in User.query.filter_by(is_active=True).all()]

    if form.validate_on_submit():
        form.populate_obj(asset)
        if form.owner_id.data == 0:
            asset.owner_id = None
        if isinstance(form.picture.data, FileStorage):
            if _allowed_picture(form.picture.data.filename):
                _delete_picture(asset.picture)
                asset.picture = _save_picture(form.picture.data)
            else:
                flash(_("Invalid picture format. Allowed: PNG, JPG, JPEG, GIF, WebP"), "warning")
        asset.updated_at = datetime.utcnow()
        db.session.commit()
        _log_audit(f"Updated asset: {asset.name}")
        flash(_("Asset updated successfully."), "success")
        return redirect(url_for("assets.view_asset", asset_id=asset.id))

    form.owner_id.data = asset.owner_id or 0
    return render_template("assets/form.html", form=form, title=_("Edit Asset"), asset=asset)


@assets_bp.route("/<int:asset_id>/delete", methods=["POST"])
@login_required
@admin_required
def delete_asset(asset_id):
    asset = Asset.query.get_or_404(asset_id)
    name = asset.name
    _delete_picture(asset.picture)
    db.session.delete(asset)
    db.session.commit()
    _log_audit_action(f"Deleted asset: {name}")
    flash(_("Asset deleted."), "success")
    return redirect(url_for("assets.list_assets"))


def _log_audit(details):
    _log_audit_action(details)


def _log_audit_action(details):
    try:
        log = AuditLog(
            user_id=current_user.id,
            action="DELETE" if "Deleted" in details else "CREATE" if "Created" in details else "UPDATE",
            resource_type="Asset",
            details=details,
            ip_address=request.remote_addr,
            user_agent=request.headers.get("User-Agent", "")[:256],
        )
        db.session.add(log)
        db.session.commit()
    except Exception:
        pass
